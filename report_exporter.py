"""
report_exporter.py — PDF & Excel Audit Report Exporter for CallIQ (Milestone 4)

Generates professional downloadable reports from evaluation results.

PDF: Uses reportlab (pure Python, no browser/LibreOffice needed).
Excel: Uses openpyxl with full formatting, colour-coded scores, charts.

Both reports include:
  - Cover/summary section
  - 5-dimension scores with colour coding
  - Compliance violations & keywords
  - Improvement suggestions
  - Full transcript
"""

import io
import datetime
from pathlib import Path

# ── Grade helper ──────────────────────────────────────────────────────────────
def grade(total):
    pct = int((total / 25) * 100)
    g = "A" if pct >= 90 else ("B" if pct >= 75 else ("C" if pct >= 60 else ("D" if pct >= 50 else "F")))
    return g, pct

def score_hex(v):
    return "34D399" if v >= 4 else ("FBBF24" if v >= 3 else "F87171")


# ══════════════════════════════════════════════════════════════════════════════
# PDF EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def build_pdf_report(result: dict, transcript: str) -> bytes:
    """
    Generate a professionally formatted PDF audit report.
    Returns raw bytes ready for Streamlit download_button.
    Requires: pip install reportlab
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak, KeepTogether
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    # ── Colours ──────────────────────────────────────────────────────────────
    DARK_BLUE  = colors.HexColor("#1F3864")
    MID_BLUE   = colors.HexColor("#2E75B6")
    ACCENT     = colors.HexColor("#FF4D6D")
    GREEN      = colors.HexColor("#34D399")
    AMBER      = colors.HexColor("#FBBF24")
    RED        = colors.HexColor("#F87171")
    LIGHT_GRAY = colors.HexColor("#F2F2F2")
    DARK_GRAY  = colors.HexColor("#595959")
    WHITE      = colors.white

    def score_color(v):
        return GREEN if v >= 4 else (AMBER if v >= 3 else RED)

    # ── Styles ────────────────────────────────────────────────────────────────
    styles = getSampleStyleSheet()
    title_style  = ParagraphStyle("title",  fontName="Helvetica-Bold",   fontSize=24, textColor=DARK_BLUE, spaceAfter=6,  alignment=TA_CENTER)
    sub_style    = ParagraphStyle("sub",    fontName="Helvetica",        fontSize=11, textColor=DARK_GRAY,  spaceAfter=4,  alignment=TA_CENTER)
    h1_style     = ParagraphStyle("h1",     fontName="Helvetica-Bold",   fontSize=14, textColor=DARK_BLUE,  spaceBefore=16, spaceAfter=6)
    h2_style     = ParagraphStyle("h2",     fontName="Helvetica-Bold",   fontSize=11, textColor=MID_BLUE,   spaceBefore=10, spaceAfter=4)
    body_style   = ParagraphStyle("body",   fontName="Helvetica",        fontSize=9,  textColor=colors.black, spaceAfter=4, leading=14)
    small_style  = ParagraphStyle("small",  fontName="Helvetica",        fontSize=8,  textColor=DARK_GRAY,  spaceAfter=2)
    mono_style   = ParagraphStyle("mono",   fontName="Courier",          fontSize=7.5,textColor=DARK_GRAY,  spaceAfter=2, leading=12)
    label_style  = ParagraphStyle("label",  fontName="Helvetica-Bold",   fontSize=8,  textColor=DARK_GRAY,  spaceAfter=1)
    badge_green  = ParagraphStyle("bg",     fontName="Helvetica-Bold",   fontSize=9,  textColor=colors.HexColor("#065F46"), alignment=TA_CENTER)
    badge_amber  = ParagraphStyle("ba",     fontName="Helvetica-Bold",   fontSize=9,  textColor=colors.HexColor("#92400E"), alignment=TA_CENTER)
    badge_red    = ParagraphStyle("br",     fontName="Helvetica-Bold",   fontSize=9,  textColor=colors.HexColor("#7F1D1D"), alignment=TA_CENTER)

    # ── Data ──────────────────────────────────────────────────────────────────
    scores   = result.get("scores", {})
    total    = result.get("total_score", 0)
    descs    = result.get("score_descriptions", {})
    sent     = result.get("customer_sentiment", "Neutral")
    escal    = result.get("escalation_risk", "Medium")
    viols    = result.get("compliance_violations", [])
    kws      = result.get("compliance_keywords", [])
    suggs    = result.get("improvement_suggestions", [])
    strengths= result.get("agent_strengths", [])
    summary  = result.get("call_summary", "")
    fname    = result.get("filename", "")
    ts       = result.get("timestamp", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    g_ltr, g_pct = grade(total)

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=letter,
                              leftMargin=0.75*inch, rightMargin=0.75*inch,
                              topMargin=0.75*inch,  bottomMargin=0.75*inch)
    story = []

    # ── Cover ──────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph("CallIQ", title_style))
    story.append(Paragraph("AI-Powered Customer Support Quality Audit Report", sub_style))
    story.append(HRFlowable(width="100%", thickness=2, color=ACCENT, spaceAfter=12))

    # Summary metrics table
    sent_color  = GREEN if sent=="Positive" else (AMBER if sent=="Neutral" else RED)
    escal_color = GREEN if escal=="Low" else (AMBER if escal=="Medium" else RED)
    grade_color = GREEN if g_ltr in ("A","B") else (AMBER if g_ltr=="C" else RED)

    meta_table = Table([
        [Paragraph("FILE", label_style), Paragraph("TIMESTAMP", label_style),
         Paragraph("SCORE", label_style), Paragraph("GRADE", label_style),
         Paragraph("SENTIMENT", label_style), Paragraph("ESCALATION", label_style)],
        [Paragraph(fname, body_style), Paragraph(ts, small_style),
         Paragraph(f"{total}/25", ParagraphStyle("sc", fontName="Helvetica-Bold", fontSize=16, textColor=ACCENT, alignment=TA_CENTER)),
         Paragraph(g_ltr, ParagraphStyle("gr", fontName="Helvetica-Bold", fontSize=16, textColor=grade_color, alignment=TA_CENTER)),
         Paragraph(sent, ParagraphStyle("se", fontName="Helvetica-Bold", fontSize=11, textColor=sent_color, alignment=TA_CENTER)),
         Paragraph(escal, ParagraphStyle("es", fontName="Helvetica-Bold", fontSize=11, textColor=escal_color, alignment=TA_CENTER))],
    ], colWidths=[1.8*inch, 1.5*inch, 0.9*inch, 0.7*inch, 1*inch, 1*inch])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#D6E4F0")),
        ("BACKGROUND", (0,1), (-1,1), colors.HexColor("#F8F8FC")),
        ("BOX",        (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
        ("INNERGRID",  (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING", (0,0),(-1,-1), 8),
        ("RIGHTPADDING",(0,0),(-1,-1), 8),
        ("ALIGN",      (0,0),(-1,-1), "CENTER"),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.15*inch))

    # Call summary
    if summary:
        story.append(Paragraph("Call Summary", h1_style))
        story.append(Paragraph(summary, body_style))
        story.append(Spacer(1, 0.08*inch))

    # ── Dimension Scores ──────────────────────────────────────────────────
    story.append(Paragraph("Dimension Scores", h1_style))
    dim_labels = [
        ("greeting_quality",      "Greeting Quality"),
        ("empathy",               "Empathy"),
        ("problem_understanding", "Problem Understanding"),
        ("resolution_clarity",    "Resolution Clarity"),
        ("professionalism",       "Professionalism"),
    ]
    dim_rows = [["Dimension", "Score", "Rating", "AI Verdict"]]
    for key, label in dim_labels:
        v = scores.get(key, 0)
        rating = "Excellent" if v==5 else ("Good" if v==4 else ("Fair" if v==3 else ("Weak" if v==2 else "Poor")))
        dim_rows.append([
            Paragraph(label, body_style),
            Paragraph(f"{v}/5", ParagraphStyle("ds", fontName="Helvetica-Bold", fontSize=10, textColor=score_color(v), alignment=TA_CENTER)),
            Paragraph(rating, small_style),
            Paragraph(descs.get(key, ""), small_style),
        ])

    dim_table = Table(dim_rows, colWidths=[1.6*inch, 0.7*inch, 0.8*inch, 3.8*inch])
    dim_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,0), DARK_BLUE),
        ("TEXTCOLOR",  (0,0),(-1,0), WHITE),
        ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0),(-1,0), 9),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.HexColor("#F8F8FC"), WHITE]),
        ("BOX",        (0,0),(-1,-1), 0.5, colors.HexColor("#CCCCCC")),
        ("INNERGRID",  (0,0),(-1,-1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0,0),(-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1), 8),
        ("RIGHTPADDING",(0,0),(-1,-1),8),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
    ]))
    story.append(dim_table)
    story.append(Spacer(1, 0.12*inch))

    # ── Agent Strengths ────────────────────────────────────────────────────
    if strengths:
        story.append(Paragraph("Agent Strengths", h1_style))
        for s in strengths:
            story.append(Paragraph(f"✓  {s}", body_style))
        story.append(Spacer(1, 0.08*inch))

    # ── Compliance ────────────────────────────────────────────────────────
    story.append(Paragraph("Compliance", h1_style))
    if viols:
        story.append(Paragraph("Violations Detected", h2_style))
        for v in viols:
            story.append(Paragraph(f"⚠  {v}", ParagraphStyle("viol", fontName="Helvetica", fontSize=9, textColor=RED, spaceAfter=3)))
    else:
        story.append(Paragraph("✓  No compliance violations detected.", ParagraphStyle("ok", fontName="Helvetica", fontSize=9, textColor=GREEN, spaceAfter=4)))

    if kws:
        story.append(Paragraph(f"Compliance Keywords Detected:  {',  '.join(kws)}", small_style))
    story.append(Spacer(1, 0.08*inch))

    # ── Improvement Suggestions ────────────────────────────────────────────
    if suggs:
        story.append(Paragraph("Improvement Suggestions", h1_style))
        for i, s in enumerate(suggs):
            if isinstance(s, dict):
                pri   = s.get("priority","Medium")
                pri_c = RED if pri=="High" else (AMBER if pri=="Medium" else GREEN)
                block = [
                    [Paragraph(f"[{i+1}]  {s.get('title','')}", ParagraphStyle("sh", fontName="Helvetica-Bold", fontSize=9, textColor=DARK_BLUE)),
                     Paragraph(f"{pri} Priority  ·  {s.get('category','')}", ParagraphStyle("sp", fontName="Helvetica-Bold", fontSize=8, textColor=pri_c, alignment=TA_RIGHT))],
                    [Paragraph(f"Issue:  {s.get('issue','')}", small_style), ""],
                    [Paragraph(f"Action: {s.get('action','')}", small_style), ""],
                ]
                if s.get("example"):
                    block.append([Paragraph(f"Example: {s.get('example','')}", ParagraphStyle("ex", fontName="Courier", fontSize=8, textColor=DARK_GRAY)), ""])
                t = Table(block, colWidths=[5*inch, 1.9*inch])
                t.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#EEF4FB")),
                    ("SPAN",      (0,1),(1,1)),
                    ("SPAN",      (0,2),(1,2)),
                    ("SPAN",      (0,3),(1,3)) if len(block)>3 else ("", (0,0),(0,0)),
                    ("BOX",       (0,0),(-1,-1), 0.5, colors.HexColor("#CCCCCC")),
                    ("INNERGRID", (0,0),(-1,-1), 0.3, colors.HexColor("#EEEEEE")),
                    ("TOPPADDING",(0,0),(-1,-1), 5),
                    ("BOTTOMPADDING",(0,0),(-1,-1),5),
                    ("LEFTPADDING",(0,0),(-1,-1), 8),
                ]))
                story.append(KeepTogether([t, Spacer(1, 6)]))
            else:
                story.append(Paragraph(f"{i+1}.  {s}", body_style))

    # ── Transcript ────────────────────────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("Full Transcript", h1_style))
    story.append(HRFlowable(width="100%", thickness=1, color=MID_BLUE, spaceAfter=8))
    for line in (transcript or "").splitlines():
        if line.strip():
            story.append(Paragraph(line, mono_style))

    # ── Footer note ────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.2*inch))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CCCCCC"), spaceAfter=6))
    story.append(Paragraph("Generated by CallIQ · Vidzai Digital · AI evaluations are assistive tools — not a substitute for human judgment.", small_style))

    doc.build(story)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def build_excel_report(result: dict, transcript: str) -> bytes:
    """
    Generate a formatted Excel audit report (.xlsx).
    Returns raw bytes ready for Streamlit download_button.
    Requires: pip install openpyxl
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    wb = Workbook()

    def fill(hex_c):
        return PatternFill("solid", start_color=hex_c, end_color=hex_c)

    def font(bold=False, color="000000", size=10, italic=False):
        return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

    def thin():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    scores   = result.get("scores", {})
    total    = result.get("total_score", 0)
    descs    = result.get("score_descriptions", {})
    sent     = result.get("customer_sentiment", "Neutral")
    escal    = result.get("escalation_risk", "Medium")
    viols    = result.get("compliance_violations", [])
    kws      = result.get("compliance_keywords", [])
    suggs    = result.get("improvement_suggestions", [])
    strengths= result.get("agent_strengths", [])
    summary  = result.get("call_summary", "")
    fname    = result.get("filename", "")
    ts       = result.get("timestamp", "")
    g_ltr, g_pct = grade(total)

    dim_labels = [
        ("greeting_quality",      "Greeting Quality"),
        ("empathy",               "Empathy"),
        ("problem_understanding", "Problem Understanding"),
        ("resolution_clarity",    "Resolution Clarity"),
        ("professionalism",       "Professionalism"),
    ]

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Audit Summary"

    # Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = "CallIQ — Customer Support Quality Audit Report"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = fill("1F3864")
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = f"File: {fname}   |   Date: {ts}"
    ws["A2"].font = Font(name="Arial", color="FFFFFF", size=10, italic=True)
    ws["A2"].fill = fill("2E75B6")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

    # Key metrics
    metrics = [
        ("Total Score", f"{total}/25", "FF4D6D"),
        ("Grade", g_ltr, "FF4D6D"),
        ("Percentage", f"{g_pct}%", "FF4D6D"),
        ("Customer Sentiment", sent, "34D399" if sent=="Positive" else ("FBBF24" if sent=="Neutral" else "F87171")),
        ("Escalation Risk", escal, "34D399" if escal=="Low" else ("FBBF24" if escal=="Medium" else "F87171")),
        ("Violations", str(len(viols)), "F87171" if viols else "34D399"),
    ]
    for col, (label, value, vc) in enumerate(metrics, 1):
        ws.cell(row=4, column=col, value=label).font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        ws.cell(row=4, column=col).fill = fill("2E75B6")
        ws.cell(row=4, column=col).alignment = center()
        ws.cell(row=4, column=col).border = thin()
        ws.cell(row=5, column=col, value=value).font = Font(name="Arial", bold=True, color=vc, size=14)
        ws.cell(row=5, column=col).alignment = center()
        ws.cell(row=5, column=col).border = thin()
        ws.cell(row=5, column=col).fill = fill("F2F2F2")
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 28

    # Call summary
    ws.merge_cells("A7:G7")
    ws["A7"] = "Call Summary"
    ws["A7"].font = Font(name="Arial", bold=True, color="1F3864", size=11)
    ws.row_dimensions[7].height = 18
    ws.merge_cells("A8:G8")
    ws["A8"] = summary
    ws["A8"].font = Font(name="Arial", size=9)
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[8].height = 40

    # Dimension scores table
    ws.merge_cells("A10:G10")
    ws["A10"] = "Dimension Scores"
    ws["A10"].font = Font(name="Arial", bold=True, color="1F3864", size=11)
    ws.row_dimensions[10].height = 18

    hdr_row = ["Dimension", "Score", "/ 5", "Rating", "AI Verdict", "", ""]
    for c, h in enumerate(hdr_row, 1):
        cell = ws.cell(row=11, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill = fill("1F3864")
        cell.alignment = center()
        cell.border = thin()
    ws.row_dimensions[11].height = 20

    for r, (key, label) in enumerate(dim_labels, 12):
        v = scores.get(key, 0)
        vc = score_hex(v)
        rating = "Excellent" if v==5 else ("Good" if v==4 else ("Fair" if v==3 else ("Weak" if v==2 else "Poor")))
        bg = "E2EFDA" if v >= 4 else ("FFF2CC" if v >= 3 else "FDECEA")
        row_data = [label, v, "/5", rating, descs.get(key,""), "", ""]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin()
            cell.font = Font(name="Arial", bold=(c==2), color=vc if c==2 else "000000", size=9)
            cell.fill = fill(bg if c <= 4 else "FFFFFF")
            cell.alignment = center() if c <= 4 else left()
        ws.row_dimensions[r].height = 22

    ws.merge_cells("E12:G16")
    for r in range(12, 17):
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")

    # Agent strengths
    ws.merge_cells("A18:G18")
    ws["A18"] = "Agent Strengths"
    ws["A18"].font = Font(name="Arial", bold=True, color="1F3864", size=11)
    ws.row_dimensions[18].height = 18
    for i, s in enumerate(strengths, 19):
        ws.merge_cells(f"A{i}:G{i}")
        ws.cell(row=i, column=1, value=f"✓  {s}").font = Font(name="Arial", size=9, color="065F46")
        ws.cell(row=i, column=1).fill = fill("E2EFDA")
        ws.cell(row=i, column=1).border = thin()
        ws.row_dimensions[i].height = 18

    # Column widths
    for col, w in zip("ABCDEFG", [26, 7, 5, 12, 38, 5, 5]):
        ws.column_dimensions[col].width = w

    # ── Sheet 2: Scores Chart ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Score Chart")
    ws2["A1"] = "Dimension"
    ws2["B1"] = "Score"
    ws2["A1"].font = Font(name="Arial", bold=True)
    ws2["B1"].font = Font(name="Arial", bold=True)
    for i, (key, label) in enumerate(dim_labels, 2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=scores.get(key, 0))

    chart = BarChart()
    chart.type  = "bar"
    chart.title = f"Dimension Scores — {fname}"
    chart.y_axis.title = "Score (0–5)"
    chart.x_axis.title = "Dimension"
    chart.style = 10
    chart.width = 18
    chart.height = 12
    data = Reference(ws2, min_col=2, min_row=1, max_row=6)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "D2")

    # ── Sheet 3: Compliance & Suggestions ────────────────────────────────────
    ws3 = wb.create_sheet("Compliance & Suggestions")
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "Compliance Violations"
    ws3["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    ws3["A1"].fill = fill("1F3864")
    ws3["A1"].alignment = center()
    ws3.row_dimensions[1].height = 22
    if viols:
        for i, v in enumerate(viols, 2):
            ws3.merge_cells(f"A{i}:E{i}")
            ws3.cell(row=i, column=1, value=f"⚠  {v}").font = Font(name="Arial", size=9, color="7F1D1D")
            ws3.cell(row=i, column=1).fill = fill("FDECEA")
            ws3.cell(row=i, column=1).border = thin()
            ws3.row_dimensions[i].height = 18
    else:
        ws3.merge_cells("A2:E2")
        ws3["A2"] = "✓  No compliance violations detected."
        ws3["A2"].font = Font(name="Arial", size=9, color="065F46")
        ws3["A2"].fill = fill("E2EFDA")
        ws3["A2"].border = thin()

    r = len(viols) + 4
    ws3.merge_cells(f"A{r}:E{r}")
    ws3.cell(row=r, column=1, value="Compliance Keywords Detected").font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    ws3.cell(row=r, column=1).fill = fill("1F3864")
    ws3.cell(row=r, column=1).alignment = center()
    ws3.row_dimensions[r].height = 22
    r += 1
    ws3.merge_cells(f"A{r}:E{r}")
    ws3.cell(row=r, column=1, value=",  ".join(kws) if kws else "None").font = Font(name="Arial", size=9)
    ws3.cell(row=r, column=1).border = thin()
    ws3.row_dimensions[r].height = 18

    r += 2
    ws3.merge_cells(f"A{r}:E{r}")
    ws3.cell(row=r, column=1, value="Improvement Suggestions").font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    ws3.cell(row=r, column=1).fill = fill("1F3864")
    ws3.cell(row=r, column=1).alignment = center()
    ws3.row_dimensions[r].height = 22
    r += 1

    for col, h in enumerate(["#","Priority","Category","Issue","Action"], 1):
        ws3.cell(row=r, column=col, value=h).font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        ws3.cell(row=r, column=col).fill = fill("2E75B6")
        ws3.cell(row=r, column=col).alignment = center()
        ws3.cell(row=r, column=col).border = thin()
    ws3.row_dimensions[r].height = 20
    r += 1

    for i, s in enumerate(suggs, 1):
        if isinstance(s, dict):
            pri = s.get("priority","Medium")
            vc  = "7F1D1D" if pri=="High" else ("92400E" if pri=="Medium" else "065F46")
            bg  = "FDECEA" if pri=="High" else ("FFF2CC" if pri=="Medium" else "E2EFDA")
            row_vals = [i, pri, s.get("category",""), s.get("issue",""), s.get("action","")]
            for c, val in enumerate(row_vals, 1):
                cell = ws3.cell(row=r, column=c, value=val)
                cell.font = Font(name="Arial", size=9, color=vc if c==2 else "000000")
                cell.fill = fill(bg if c <= 3 else "FFFFFF")
                cell.alignment = center() if c <= 3 else left()
                cell.border = thin()
            ws3.row_dimensions[r].height = 22
            r += 1

    for col, w in zip("ABCDE", [5, 12, 18, 38, 38]):
        ws3.column_dimensions[col].width = w

    # ── Sheet 4: Full Transcript ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Full Transcript")
    ws4.merge_cells("A1:B1")
    ws4["A1"] = "Full Transcript"
    ws4["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    ws4["A1"].fill = fill("1F3864")
    ws4["A1"].alignment = center()
    ws4.row_dimensions[1].height = 24
    ws4.column_dimensions["A"].width = 12
    ws4.column_dimensions["B"].width = 90
    for i, line in enumerate((transcript or "").splitlines(), 2):
        if line.strip():
            ws4.cell(row=i, column=1, value=i-1).font = Font(name="Arial", size=8, color="999999")
            ws4.cell(row=i, column=2, value=line).font = Font(name="Courier New", size=9)
            ws4.cell(row=i, column=2).alignment = Alignment(wrap_text=False, vertical="center")
            ws4.row_dimensions[i].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
